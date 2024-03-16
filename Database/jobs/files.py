import csv
import os
from datetime import datetime, timedelta
from pathlib import Path, PosixPath
from shutil import copyfile
from typing import Any, Optional
from types_db import TaskData
#from jobs.ftp import ftp_conn
from openpyxl import Workbook, load_workbook
#from fromsecret import get_ftp_secret


#class UploadFiles:
#    """Carica files su di una risorsa remota.
#
#    Esempio:
#
#    ```
#    upload = UploadFiles(
#        auth_label: str     # la share dove caricare i file
#        path_label: str (default 'path')    # a qualche chiave trovo i path nel TaskData in ingresso
#        dest_prefix: [str | None]   # un eventuale path remoto dove caricare i file
#        )
#    files_creati = upload({"data": [{"Path": "/file/to/upload.csv" ....
#
#    il run può prendere altri due paramteri opzionali :
#
#    dest_prefix: str    # un ulteriore parte di path da appendere al dest_prefix configurato
#    name_postfix: str   # una stringa da aggiungere in coda al nome del file
#    ```
#    """
#    """Carica files su di una risorsa remota.
#
#    Esempio:
#
#    ```
#    upload = UploadFiles(
#        auth_label: str     # la share dove caricare i file
#        path_label: str (default 'path')    # a qualche chiave trovo i path nel TaskData in ingresso
#        dest_prefix: [str | None]   # un eventuale path remoto dove caricare i file
#        )
#    files_creati = upload({"data": [{"Path": "/file/to/upload.csv" ....
#
#    il run può prendere altri due paramteri opzionali :
#
#    dest_prefix: str    # un ulteriore parte di path da appendere al dest_prefix configurato
#    name_postfix: str   # una stringa da aggiungere in coda al nome del file
#    ```
#    """
#
#    def __init__(
#        self,
#        auth_label: str,
#        path_label: str = "path",
#        remotename_label: str | None = None,
#        dest_prefix: str | None = None,
#        **kwargs: Any,
#    ):
#        self.auth_label = auth_label
#        self.path_label = path_label
#        self.dest_prefix = dest_prefix or ""
#        self.remotename_label = remotename_label
#        super().__init__(**kwargs)
#
#    def run(  # type: ignore
#        self, rows: TaskData, dest_prefix: Optional[str] = None, name_postfix: Optional[str] = None
#    ) -> TaskData:
#        with self.log.start_action(self.name), self.log.timed(self.name):
#            output = []
#            with ftp_conn(get_ftp_secret(self.auth_label)) as conn:
#                for row in rows["data"]:
#                    path = row[self.path_label]
#                    prefix = self.dest_prefix if not dest_prefix else f"{self.dest_prefix}/{dest_prefix}".lstrip("/")
#                    if self.remotename_label:
#                        file_name = PosixPath(row[self.remotename_label])
#                    else:
#                        file_name = PosixPath(path.name)
#                    if name_postfix:
#                        file_name = PosixPath(f"{file_name}{name_postfix}")
#                    remote_path = conn.store(PosixPath(path), prefix / file_name)
#                    output.append({"path": remote_path})
#
#            return {"data": output, "errors": [], "meta": {"isEmpty": len(output) == 0}}
#
def _clean_options(type: str, options: dict[str, str]) -> dict[str, str]:
    permitted_options = {
        "csv": ["delimiter", "quotechar", "quoting", "dialect"],
        "excel": ["sheet_label", "custom_headers"],
    }

    keys = set(permitted_options[type]).intersection(set(options.keys()))
    return {k: options[k] for k in keys}


class ABCWrite:
    def __init__(
        self,
        filename: Optional[str] = None,
        filename_label: Optional[str] = None,
        filename_meta: Optional[tuple[str, ...]] = None,
        folder_labels: tuple[str, ...] = (),
        options: dict[str, Any] = {},
        errors: bool = False,
        env_in_filename: bool = False,
        date_in_filename: Optional[str] = None,
        headless: bool = False,
        encoding: str = "utf-8",
        **kwargs: Any,
    ) -> None:
        self.encoding = encoding
        self.folder_labels = folder_labels
        self.filename_label = filename_label
        self.filename_meta = filename_meta
        self.filename = filename
        self.options = options
        self.errors = errors
        self.env_in_filename = env_in_filename
        self.date_fname_pre = date_in_filename.endswith("_") if date_in_filename else None
        self.date_in_filename = date_in_filename.rstrip("_") if date_in_filename else None
        self.headless = headless
        self.writers: dict[Path, Any] = {}
        self.target_files: list[Any] = []
        super().__init__(**kwargs)

    def infere_filepath(  # noqa
        self, row: dict[str, Any], tsuffix: str, meta: dict[str, Any], suffix: Optional[str], filename: Optional[str]
    ) -> Path:

        if all([fn is None for fn in (self.filename, self.filename_label, self.filename_meta, filename)]):
            raise ValueError("Definisci un modo di risolvere il nome del file.")

        if len([o for o in (self.filename, self.filename_label, self.filename_meta, filename) if o is not None]) > 1:
            raise ValueError("Il nome del file deve essere definito in un unico modo.")

        if filename is not None:
            pass
        elif self.filename_label:
            filename = row[self.filename_label]
        else:
            filename = self.filename

        if self.filename_meta:
            filename = "_".join(meta[s] for s in self.filename_meta)

        fname_date = f"{datetime.now().strftime(self.date_in_filename)}" if self.date_in_filename else ""
        env_prefix = f"{os.getenv('DEPLOY_ENV')}" if self.env_in_filename else ""

        if self.date_fname_pre:
            filename = f"{fname_date}_{env_prefix}_{filename}"

        if filename.endswith(f".{tsuffix}"):  # type: ignore
            filename = filename[0 : -len(tsuffix) - 1]  # type: ignore

        if not self.date_fname_pre:
            filename = f"{env_prefix}_{filename}_{fname_date}"

        filename = filename.strip("_")  # type: ignore

        if suffix:
            filename = f"{filename}{suffix}"

        if not filename.endswith(f".{tsuffix}"):
            filename = f"{filename}.{tsuffix}"

        path = Path(
            Path("/".join([str(row[label]).replace("\\", "/") for label in self.folder_labels])),
        )
        path.mkdir(exist_ok=True, parents=True)
        return Path(path / filename)

    def _write(
        self, data: TaskData, template: str = "", suffix: Optional[str] = None, filename: Optional[str] = None
    ) -> TaskData:
        raise NotImplementedError


class ABCFileWrite(ABCWrite):
    def __init__(self, fields: list[str], *args: Any, **kwargs: Any) -> None:
        self.fieldnames = fields
        super().__init__(*args, **kwargs)

    def run(  # type: ignore
        self, data: TaskData, template: str = "", suffix: Optional[str] = None, filename: Optional[str] = None
    ) -> TaskData:
        return self._write(data, template, suffix, filename)


class ABCTemplateWrite(ABCWrite):
    def __init__(self, mapping: dict[str, str], *args: Any, **kwargs: Any) -> None:
        self.mapping = mapping
        super().__init__(*args, **kwargs)

    def run(  # type: ignore
        self, data: TaskData, template: str = "", suffix: Optional[str] = None, filename: Optional[str] = None
    ) -> TaskData:
        return self._write(data, template, suffix, filename)


class WriteCsv(ABCFileWrite):
    """Scrive i dati di un TaskData su di un file CSV.

    ```
    write = WriteCsv(
        fields: # i nomi delle colonne da salvare
        filename: # il nome del file, oppure
        filename_label: # il nome della colonna dati da cui prendere il nome del file, oppure
        filename_meta: # il nome del campo dei meta da cui prendere il nome del file,
        folder_labels: # gli eventuali nomi delle colonne dati da cui prendere le directory in cui salvare i file,
        options: # opzioni specifiche per il csv (delimiter, quotechar, lineterminator)
        errors: # se True, scrive gli errors del TaskData invece che i data,
        env_in_filename: # se True, prepende il nome del file con l'env in cui il flow ha girato,
        date_in_filename: # se presente, è una stringa che specifica il formato data con cui prependere il nome
        headless: # se True, considera il file headless
    )
    filespath = write(dati) # torna un TaskData coi path dei file creati
    ```
    """

    def _write(
        self, data: TaskData, template: str = "", suffix: Optional[str] = None
        , filename: Optional[str] = None
        ,platform: str = "Windows"
    ) -> TaskData:
        options = _clean_options(type="csv", options=self.options)
        for row in data["data" if not self.errors else "errors"]:  # type: ignore
            path = self.infere_filepath(row, "csv", data["meta"], suffix, filename)
            if path not in self.writers:
                if platform == "Windows":
                    f = open(path.as_posix(), "w", encoding=self.encoding,newline='')
                else:
                    f = open(path.as_posix(), "w", encoding=self.encoding)
                writer = csv.DictWriter(f, fieldnames=self.fieldnames, extrasaction="ignore", **options)  # type: ignore
                if not self.headless:
                    writer.writeheader()
                self.writers[path] = writer
                self.target_files.append(f)
            else:
                writer = self.writers[path]

            writer.writerow(row)

        output = []
        for opened_file in self.target_files:
            output.append({"path": Path(opened_file.name)
                           ,"basename": os.path.basename(opened_file.name)
                           })
            opened_file.close()
        return {"data": output, "errors": [], "meta": data["meta"]}


class WriteExcel(ABCFileWrite):
    """Scrive dati su di un file Excel.

    ```
    write = WriteExcel(
        fields: # i nomi delle colonne da salvare
        filename: # il nome del file, oppure
        filename_label: # il nome della colonna dati da cui prendere il nome del file, oppure
        filename_meta: # il nome del campo dei meta da cui prendere il nome del file,
        folder_labels: # gli eventuali nomi delle colonne dati da cui prendere le directory in cui salvare i file,
        options: # opzioni specifiche per l'excel (custom_headers)
        errors: # se True, scrive gli errors del TaskData invece che i data,
        env_in_filename: # se True, prepende il nome del file con l'env in cui il flow ha girato,
        date_in_filename: # se presente, è una stringa che specifica il formato data con cui prependere il nome
        headless: # se True, considera il file headless
        )
    filespath = write(dati) # torna un TaskData coi path dei file creati

    # dettaglio opzioni specifiche :
    #   se specifico il nome sheet preso a runtime poi posso associare diverse colonne
    #   da valorizzare a seconda del foglio in cui scrivo
        custom_headers = {
            "sheetname1":["colonna1","colonna2",...]
            "sheetname2":["colonna3","colonna4",...]
        }
    ```
    """

    def _order_row(
        self,
        row: dict[str, Any],
        fields: list[str],
    ) -> dict[str, Any]:
        row_: dict[str, Any] = {}
        for field in fields:
            if field in row:
                row_[field] = row[field]
            else:
                raise ValueError(f" il campo <{field}> non è presente nella riga con colonne {list(row.keys())}")
        return row_

    def _infere_headers(
        self, sheet: str
    ) -> list[str,]:
        options = _clean_options("excel", self.options)
        headers_ = options.get("custom_headers", None)
        sheet_header = headers_[sheet] if headers_ and headers_[sheet] else []  # type: ignore
        if sheet_header:
            return sheet_header  # type: ignore
        if self.fieldnames:
            return self.fieldnames
        return []

    def infere_sheet_name(self, row: dict[str, Any]) -> tuple[dict[str, Any], str]:
        options = _clean_options("excel", self.options)
        sheet_label = options.get("sheet_label", None)
        sheet_ = str(row[sheet_label]) if sheet_label else "Sheet"
        fieldnames_ = self._infere_headers(sheet_)
        row_ = self._order_row(row=row, fields=fieldnames_)
        return row_, sheet_

    def _write(
        self, data: TaskData, template: str = "", suffix: Optional[str] = None, filename: Optional[str] = None
    ) -> TaskData:
        for row in data["data" if not self.errors else "errors"]:  # type: ignore
            path = self.infere_filepath(row, "xlsx", data["meta"], suffix, filename)
            filepath = path.as_posix()
            row_, sheet = self.infere_sheet_name(row)
            _headers = self._infere_headers(sheet)
            if path not in self.writers:
                wb = Workbook()
                std = wb["Sheet"]
                wb.remove(std)
                if sheet not in wb.sheetnames:
                    wb.create_sheet(sheet)
                    writer = wb[sheet]

                    if not self.headless and _headers:
                        writer.append(_headers)
                writer.append(tuple(row_.values()))

                self.writers[path] = wb
                self.target_files.append((wb, filepath))
            else:
                wb = self.writers[path]
                if sheet not in wb.sheetnames:
                    wb.create_sheet(sheet)
                    writer = wb[sheet]

                    if not self.headless and _headers:
                        writer.append(_headers)
                else:
                    writer = wb[sheet]
                writer.append(tuple(row_.values()))
        output = []
        for wb, filepath in self.target_files:
            output.append({"path": Path(filepath)})
            wb.save(filepath)
        return {"data": output, "errors": [], "meta": {"isEmpty": len(output) == 0}}


class WriteExcelTemplate(ABCTemplateWrite):
    """Scrive dati su di un file Excel preesistente in celle mappate.

    Esempio:

    Per scrivere in una copia di file excel esistente '10' nella cella 'A1'

    ```
    write = WriteExcelTemplate(
        mapping={"K1": "A11"}
    )
    data = {"data": [{"K1": 10}]}
    filepath = write(data, template)
    ```
    """

    def _writer(self, template: Path, path: Path) -> Workbook:
        if path not in self.writers:
            filename = path.as_posix()
            copyfile(template, filename)
            writer = load_workbook(filename)

            self.writers[path] = writer
            self.target_files.append((writer, filename))
        else:
            writer = self.writers[path]
        return writer

    def _asset_path(self, template: str) -> Path:
        tenancy = os.getenv("TENANCY")
        if tenancy is not None:
            return Path("/app", ".assets", tenancy, template)
        else:
            raise ValueError("You must export TENANCY value")

    def _write(
        self, data: TaskData, template: str = "", suffix: Optional[str] = None, filename: Optional[str] = None
    ) -> TaskData:
        for row in data["data" if not self.errors else "errors"]:  # type: ignore
            path = self.infere_filepath(row, "xlsx", data["meta"], suffix, filename)
            writer = self._writer(self._asset_path(template), path)
            ws = writer.active
            for k, v in row.items():
                ws[self.mapping[k]] = v

        output = []
        for writer, filepath in self.target_files:
            output.append({"path": Path(filepath)})
            writer.save(filepath)

        return {"data": output, "errors": [], "meta": {"isEmpty": len(output) == 0}}

